"""解析 excel 目录下的原子品类关键词规则数据。"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]
from tqdm import tqdm  # type: ignore[import-untyped]

# 关键词单元格内分隔符：逗号、顿号
KEYWORD_SEP = ",，、"


def _split_keywords(cell_value: str | None) -> list[str]:
    """将单元格内容按分隔符拆成关键词列表，空或 None 返回 []。"""
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    for sep in KEYWORD_SEP:
        s = s.replace(sep, "\x00")
    return [k.strip() for k in s.split("\x00") if k.strip()]


def _parse_keyword_group(cell_value: str | None) -> list[str]:
    """解析单个关键词组单元格 → 关键词列表（用于关键词组5、一定不包含）。"""
    return _split_keywords(cell_value)


def _parse_must_contain_group(cell_value: str | None) -> list[str]:
    """解析「需同时包含」的关键词组单元格 → 关键词列表（组内全部出现才满足）。"""
    return _split_keywords(cell_value)


@dataclass
class CategoryRule:
    """单条品类匹配规则。"""

    level1_category: str = ""  # 一级原子品类
    category_code: str | int = ""  # 品类编码
    atomic_category: str = ""  # 原子品类
    sequence_no: int = 0  # 序号
    # 关键词组1-4：单元格需同时包含以下关键词（可断开）
    keyword_group_1: list[str] = field(default_factory=list)
    keyword_group_2: list[str] = field(default_factory=list)
    keyword_group_3: list[str] = field(default_factory=list)
    keyword_group_4: list[str] = field(default_factory=list)
    # 关键词组5：单元格可任意包含其中一个关键词即可
    keyword_group_5: list[str] = field(default_factory=list)
    # 一定不包含（且）
    must_not_contain: list[str] = field(default_factory=list)


@dataclass
class RuleSheetMeta:
    """表头元数据：第1行判断逻辑、第2行字段解释。"""

    logic_descriptions: list[str] = field(default_factory=list)  # 判断逻辑，每列对应
    field_descriptions: list[str] = field(default_factory=list)  # 字段解释，每列对应


@dataclass
class VerifiedBrand:
    """已校验过的品牌及其对应原子品类（与「校验过的品牌对应原子品类.xlsx」列一致）。"""

    brand_code: str | int = ""  # 品牌编码
    brand_name: str = ""  # 品牌名称
    brand_keywords: str = ""  # 品牌关键词（，表示同时包含）
    atomic_category: str = ""  # 原子品类


def load_verified_brands(excel_path: str | Path) -> list[VerifiedBrand]:
    """
    解析「校验过的品牌对应原子品类」Excel，第 1 行为表头，第 2 行起为数据。
    表头列：品牌编码、品牌名称、品牌关键词（，表示同时包含）、原子品类。
    """
    path = Path(excel_path)
    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2:
        wb.close()
        return []

    header = [_cell_value(ws.cell(1, c)) for c in range(1, max_col + 1)]
    col_code = _find_column(header, ("品牌编码",))
    col_name = _find_column(header, ("品牌名称",))
    col_keywords = _find_column(header, ("品牌关键词（，表示同时包含）",))
    col_atomic = _find_column(header, ("原子品类",))
    if col_name is None or col_atomic is None:
        wb.close()
        return []

    result: list[VerifiedBrand] = []
    for r in range(2, max_row + 1):
        brand_name = _cell_value(ws.cell(r, col_name + 1))
        if not brand_name:
            continue
        code_cell = ws.cell(r, col_code + 1).value if col_code is not None else ""
        keywords = _cell_value(ws.cell(r, col_keywords + 1)) if col_keywords is not None else ""
        atomic = _cell_value(ws.cell(r, col_atomic + 1))
        result.append(
            VerifiedBrand(
                brand_code=code_cell if code_cell != "" else "",
                brand_name=brand_name,
                brand_keywords=keywords,
                atomic_category=atomic,
            )
        )
    wb.close()
    return result


def _find_column(header: list[str], names: tuple[str, ...]) -> int | None:
    """在表头中查找任一名称所在的列索引（0-based），未找到返回 None。"""
    for i, h in enumerate(header):
        if h and h.strip() in names:
            return i
    return None


def load_rules(excel_path: str | Path) -> tuple[RuleSheetMeta, list[CategoryRule]]:
    """
    从 Excel 解析原子品类关键词规则。

    - 第 1 行：对应字段的判断逻辑
    - 第 2 行：对应字段解释
    - 第 3 行起：规则数据

    返回 (表头元数据, 规则列表)。
    """

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("工作簿中没有活动表")

    max_col = ws.max_column
    max_row = ws.max_row

    # 第1行：判断逻辑
    logic_row = [_cell_value(ws.cell(1, c)) for c in range(1, max_col + 1)]
    # 第2行：字段解释
    field_row = [_cell_value(ws.cell(2, c)) for c in range(1, max_col + 1)]

    meta = RuleSheetMeta(logic_descriptions=logic_row, field_descriptions=field_row)

    rules: list[CategoryRule] = []
    row_range = range(3, max_row + 1)
    for r in tqdm(row_range, desc="解析规则", unit="行"):
        row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        rule = _row_to_rule(row, logic_row, field_row)
        rules.append(rule)

    wb.close()
    return meta, rules


def match_rule(text: str, rule: CategoryRule) -> bool:
    """
    按判断逻辑校验：门店文本是否满足该规则。

    - 关键词组1-4：至少有一组非空且该组内所有关键词均在文本中出现（需同时包含）。
    - 关键词组5：若非空，则至少包含其中一个关键词即可。
    - 一定不包含：文本中不能出现其中任一关键词。
    """
    t = text.strip()
    if not t:
        return False
    # 一定不包含：任一出现则不匹配
    for kw in rule.must_not_contain:
        if kw and kw in t:
            return False
    # 关键词组5：非空时至少包含一个
    if rule.keyword_group_5:
        if not any(kw and kw in t for kw in rule.keyword_group_5):
            return False
    # 关键词组1-4：若有任一组非空，则至少有一组满足「组内全部关键词均在文本中」
    groups = [
        rule.keyword_group_1,
        rule.keyword_group_2,
        rule.keyword_group_3,
        rule.keyword_group_4,
    ]
    non_empty_groups = [g for g in groups if g]
    if non_empty_groups:
        if not any(all(kw in t for kw in g) for g in non_empty_groups):
            return False
    # 至少需要有关键词条件（1-4 或 5 至少一组非空），否则视为无效规则
    if not non_empty_groups and not rule.keyword_group_5:
        return False
    return True


def match_store(text: str, rules: list[CategoryRule]) -> list[CategoryRule]:
    """对门店文本循环所有规则，返回匹配的规则列表（即对应的原子分类）。"""
    return [r for r in rules if match_rule(text, r)]


def text_similarity(text_a: str, text_b: str) -> float:
    """
    计算两段文本的相似度，返回值在 [0, 1]。
    当前为空实现，后续可替换为编辑距离、余弦相似度等。
    """
    # 空实现：始终返回 0，不判定为已知品牌
    _ = text_a
    _ = text_b
    return 0.0


def match_by_similarity(
    store_text: str,
    verified_brands: list[VerifiedBrand],
    threshold: float = 0.97,
) -> list[CategoryRule]:
    """
    与已校验品牌做相似度比对，若存在相似度 >= threshold 的品牌，则返回其原子品类（单条规则）。
    否则返回空列表。
    """
    store_text = store_text.strip()
    if not store_text or not verified_brands:
        return []

    best_score = -1.0
    best_brand: VerifiedBrand | None = None
    for vb in verified_brands:
        if not vb.brand_name:
            continue
        score = text_similarity(store_text, vb.brand_name)
        if score >= threshold and score > best_score:
            best_score = score
            best_brand = vb

    if best_brand is None or best_score < threshold:
        return []

    return [
        CategoryRule(
            level1_category="",
            category_code=best_brand.brand_code,
            atomic_category=best_brand.atomic_category,
        )
    ]


def _cell_value(cell: Any) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def _row_to_rule(row: list[Any], logic_row: list[str], field_row: list[str]) -> CategoryRule:
    """将一行数据转为 CategoryRule。列索引 0-based：0 一级原子品类, 1 品类编码, 2 原子品类, 3 序号, 4-7 关键词组1-4, 8 关键词组5, 9 一定不包含。"""
    n = len(row)

    def cell(i: int) -> Any:
        return row[i] if i < n else None

    def str_cell(i: int) -> str:
        v = cell(i)
        return "" if v is None else str(v).strip()

    def int_cell(i: int) -> int:
        v = cell(i)
        if v is None:
            return 0
        if isinstance(v, int):
            return v
        try:
            return int(float(str(v).strip()))
        except (ValueError, TypeError):
            return 0

    return CategoryRule(
        level1_category=str_cell(0),
        category_code=cell(1) if cell(1) is not None else "",  # 可能是 490 或 "411、412"
        atomic_category=str_cell(2),
        sequence_no=int_cell(3),
        keyword_group_1=_parse_must_contain_group(cell(4)),
        keyword_group_2=_parse_must_contain_group(cell(5)),
        keyword_group_3=_parse_must_contain_group(cell(6)),
        keyword_group_4=_parse_must_contain_group(cell(7)),
        keyword_group_5=_parse_keyword_group(cell(8)),
        must_not_contain=_parse_keyword_group(cell(9)),
    )
