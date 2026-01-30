"""品类匹配：从文件按行读入品类，匹配原子分类并输出到 output 目录的 Excel。"""

import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

from parser import (
    CategoryRule,
    VerifiedBrand,
    load_rules,
    load_verified_brands,
    match_by_similarity,
    match_store,
)

SIMILARITY_THRESHOLD = 0.97


def match_store_categories(
    store_text: str,
    rules: list[CategoryRule],
    verified_brands: list[VerifiedBrand],
) -> tuple[list[CategoryRule], bool]:
    """
    将门店名称/描述与已加载的规则、已校验品牌作为参数，匹配并返回 (该门店对应的原子分类规则列表, 是否由相似度匹配).
    """
    matched = match_store(store_text.strip(), rules)
    from_similarity = False
    if not matched and verified_brands:
        matched = match_by_similarity(
            store_text, verified_brands, threshold=SIMILARITY_THRESHOLD
        )
        from_similarity = bool(matched)
    return matched, from_similarity


def normalize_input_path(raw: str) -> Path:
    """
    规范化用户输入的文件路径：去首尾引号/空白；在 WSL/Linux 下将 Windows 盘符路径转为可访问路径。
    例如 'c:/Users/cyy/Desktop/文件.txt' -> /mnt/c/Users/cyy/Desktop/文件.txt
    """
    s = raw.strip().strip("\"'\"''")
    if not s:
        return Path("")
    # WSL/Linux 下：将 c:/ 或 C:\ 转为 /mnt/c/
    if os.name == "posix" and len(s) >= 2:
        m = re.match(r"^([a-zA-Z])\s*[:\\](.*)$", s)
        if m:
            drive = m.group(1).lower()
            rest = (m.group(2) or "").replace("\\", "/").strip("/")
            s = f"/mnt/{drive}/{rest}" if rest else f"/mnt/{drive}"
    return Path(s)


def read_categories_from_file(file_path: Path) -> list[str]:
    """读取文件，每行一个品类，去空行、去首尾空白。"""
    try:
        text = file_path.read_text(encoding="utf-8")
    except Exception as e:
        try:
            text = file_path.read_text(encoding="utf-8-sig")
        except Exception:
            raise RuntimeError(f"无法读取文件 {file_path}: {e}") from e
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return lines


def run_batch_match(
    categories: list[str],
    rules: list[CategoryRule],
    verified_brands: list[VerifiedBrand],
) -> list[tuple[str, str, str, str, str]]:
    """
    批量匹配。返回列表，每项为 (输入品类, 一级原子品类, 品类编码, 原子品类, 匹配方式)。
    匹配方式: "规则" | "相似度" | "未匹配"。多条匹配时取第一条。
    """
    result: list[tuple[str, str, str, str, str]] = []
    for name in categories:
        matched, from_similarity = match_store_categories(name, rules, verified_brands)
        if not matched:
            result.append((name, "", "", "", "未匹配"))
            continue
        r = matched[0]
        method = "相似度" if from_similarity else "规则"
        result.append(
            (
                name,
                r.level1_category or "",
                str(r.category_code) if r.category_code != "" else "",
                r.atomic_category or "",
                method,
            )
        )
    return result


def write_result_excel(
    rows: list[tuple[str, str, str, str, str]],
    output_path: Path,
) -> None:
    """将结果写入 Excel，未匹配行标红。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = "匹配结果"
    headers = ("输入品类", "一级原子品类", "品类编码", "原子品类", "匹配方式")
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    red_font = Font(color="FF0000")
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[4] == "未匹配":
                cell.font = red_font
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    excel_dir = base_dir / "excel"
    output_dir = base_dir / "output"
    rules_path = excel_dir / "原子品类关键词.xlsx"
    verified_path = excel_dir / "校验过的品牌对应原子品类.xlsx"

    print("解析规则与已校验品牌...")
    _, rules = load_rules(rules_path)
    verified_brands = load_verified_brands(verified_path)
    print(f"规则 {len(rules)} 条，已校验品牌 {len(verified_brands)} 条。")
    print("请拖动或输入待匹配品类文件路径（每行一个品类），输入 q 或直接关闭窗口退出。\n")

    # 若启动时带了文件路径参数（如拖动到控制台），先执行一次
    pending_path: str | None = None
    if len(sys.argv) > 1:
        pending_path = " ".join(sys.argv[1:]).strip()

    while True:
        file_path_str = pending_path if pending_path else input("文件路径: ").strip()
        pending_path = None
        if not file_path_str:
            continue
        if file_path_str.lower() in ("q", "quit", "exit"):
            print("退出。")
            break

        path = normalize_input_path(file_path_str)
        if not path or not path.exists():
            print(f"文件不存在: {path}\n")
            continue

        try:
            categories = read_categories_from_file(path)
        except Exception as e:
            print(f"读取失败: {e}\n")
            continue
        if not categories:
            print("文件中没有有效品类行。\n")
            continue

        print(f"共 {len(categories)} 条品类，正在匹配...")
        result_rows = run_batch_match(categories, rules, verified_brands)
        unmatched_count = sum(1 for r in result_rows if r[4] == "未匹配")

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"{path.stem}_匹配结果_{stamp}.xlsx"
        output_path = output_dir / out_name
        write_result_excel(result_rows, output_path)
        print(f"已写入: {output_path}")
        print(f"匹配 {len(result_rows) - unmatched_count} 条，未匹配 {unmatched_count} 条（结果表中已标红）。\n")


if __name__ == "__main__":
    main()
