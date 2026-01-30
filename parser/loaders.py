"""从 Excel 加载规则与已校验品牌。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]
from tqdm import tqdm  # type: ignore[import-untyped]

from .models import CategoryRule, RuleSheetMeta, VerifiedBrand

KEYWORD_SEP = ",，、"


def _cell_value(cell: Any) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def _find_column(header: list[str], names: tuple[str, ...]) -> int | None:
    """在表头中查找任一名称所在的列索引（0-based），未找到返回 None。"""
    for i, h in enumerate(header):
        if h and h.strip() in names:
            return i
    return None


def _split_keywords(cell_value: str | None) -> list[str]:
    """将单元格内容按分隔符拆成关键词列表，空或 None 返回 []。"""
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    for sep in KEYWORD_SEP:
        s = s.replace(sep, "\x00")
    return [k.strip() for k in s.split("\x00") if k.strip()]


def _parse_keyword_group(cell_value: str | None) -> list[str]:
    """解析单个关键词组单元格 → 关键词列表（用于关键词组5、一定不包含）。"""
    return _split_keywords(cell_value)


def _parse_must_contain_group(cell_value: str | None) -> list[str]:
    """解析「需同时包含」的关键词组单元格 → 关键词列表。"""
    return _split_keywords(cell_value)


def _row_to_rule(row: list[Any], logic_row: list[str], field_row: list[str]) -> CategoryRule:
    """将一行数据转为 CategoryRule。列索引 0-based：0 一级原子品类, 1 品类编码, 2 原子品类, 3 序号, 4-7 关键词组1-4, 8 关键词组5, 9 一定不包含。"""
    n = len(row)

    def cell(i: int) -> Any:
        return row[i] if i < n else None

    def str_cell(i: int) -> str:
        v = cell(i)
        return "" if v is None else str(v).strip()

    def int_cell(i: int) -> int:
        v = cell(i)
        if v is None:
            return 0
        if isinstance(v, int):
            return v
        try:
            return int(float(str(v).strip()))
        except (ValueError, TypeError):
            return 0

    return CategoryRule(
        level1_category=str_cell(0),
        category_code=cell(1) if cell(1) is not None else "",
        atomic_category=str_cell(2),
        sequence_no=int_cell(3),
        keyword_group_1=_parse_must_contain_group(cell(4)),
        keyword_group_2=_parse_must_contain_group(cell(5)),
        keyword_group_3=_parse_must_contain_group(cell(6)),
        keyword_group_4=_parse_must_contain_group(cell(7)),
        keyword_group_5=_parse_keyword_group(cell(8)),
        must_not_contain=_parse_keyword_group(cell(9)),
    )


def load_rules(excel_path: str | Path) -> tuple[RuleSheetMeta, list[CategoryRule]]:
    """
    从 Excel 解析原子品类关键词规则。
    第 1 行：判断逻辑；第 2 行：字段解释；第 3 行起：规则数据。
    返回 (表头元数据, 规则列表)。
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("工作簿中没有活动表")

    max_col = ws.max_column
    max_row = ws.max_row
    logic_row = [_cell_value(ws.cell(1, c)) for c in range(1, max_col + 1)]
    field_row = [_cell_value(ws.cell(2, c)) for c in range(1, max_col + 1)]
    meta = RuleSheetMeta(logic_descriptions=logic_row, field_descriptions=field_row)

    rules: list[CategoryRule] = []
    for r in tqdm(range(3, max_row + 1), desc="解析规则", unit="行"):
        row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        rules.append(_row_to_rule(row, logic_row, field_row))
    wb.close()
    return meta, rules


def load_verified_brands(excel_path: str | Path) -> list[VerifiedBrand]:
    """
    解析「校验过的品牌对应原子品类」Excel，第 1 行为表头，第 2 行起为数据。
    表头列：品牌编码、品牌名称、品牌关键词（，表示同时包含）、原子品类。
    """
    path = Path(excel_path)
    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2:
        wb.close()
        return []

    header = [_cell_value(ws.cell(1, c)) for c in range(1, max_col + 1)]
    col_code = _find_column(header, ("品牌编码",))
    col_name = _find_column(header, ("品牌名称",))
    col_keywords = _find_column(header, ("品牌关键词（，表示同时包含）",))
    col_atomic = _find_column(header, ("原子品类",))
    if col_name is None or col_atomic is None:
        wb.close()
        return []

    result: list[VerifiedBrand] = []
    for r in range(2, max_row + 1):
        brand_name = _cell_value(ws.cell(r, col_name + 1))
        if not brand_name:
            continue
        code_cell = ws.cell(r, col_code + 1).value if col_code is not None else ""
        keywords = _cell_value(ws.cell(r, col_keywords + 1)) if col_keywords is not None else ""
        atomic = _cell_value(ws.cell(r, col_atomic + 1))
        result.append(
            VerifiedBrand(
                brand_code=code_cell if code_cell != "" else "",
                brand_name=brand_name,
                brand_keywords=keywords,
                atomic_category=atomic,
            )
        )
    wb.close()
    return result
