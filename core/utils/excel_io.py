"""Excel 读写公共逻辑：只读打开、单元格取值、按行写入并可选标红。"""

from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import Any, Callable

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]


def cell_value(cell_or_value: Any) -> str:
    """支持 openpyxl Cell 或裸值（如 iter_rows values_only=True），统一为 str。"""
    v = getattr(cell_or_value, "value", cell_or_value)
    if v is None:
        return ""
    return str(v).strip()


@contextmanager
def open_excel_read(path: Path):
    """以只读、data_only 方式打开 Excel，yield (wb, ws)，退出时关闭 wb。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        yield wb, wb.active
    finally:
        wb.close()


def write_sheet(
    output_path: Path,
    sheet_title: str,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
    *,
    failed_row_predicate: Callable[[tuple[Any, ...]], bool] | None = None,
    red_font_hex: str = "FF0000",
) -> None:
    """写表头与数据行；若 failed_row_predicate(row) 为 True，该行单元格标红。父目录会自动创建。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = sheet_title
    red_font = Font(color=red_font_hex)
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if failed_row_predicate and failed_row_predicate(row_data):
                cell.font = red_font
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
