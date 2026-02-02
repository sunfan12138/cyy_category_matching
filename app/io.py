"""文件读写：品类列表读取、匹配结果写入 Excel。"""

from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

# 匹配成功：关键词匹配到、相似度匹配到、搜索大模型后匹配到；其余为匹配失败
MATCH_SUCCESS_METHODS = ("规则", "相似度", "搜索后匹配")

# 7 列：(输入品类, 一级原子品类, 品类编码, 原子品类, 匹配方式, 相似度匹配结果, 大模型描述)
ResultRow = tuple[str, str, str, str, str, str, str]

HEADERS = (
    "输入品类",
    "一级原子品类",
    "品类编码",
    "原子品类",
    "匹配方式",
    "相似度匹配结果",
    "大模型描述",
)


def read_categories_from_file(file_path: Path) -> list[str]:
    """读取文件，每行一个品类，去空行、去首尾空白。"""
    try:
        text = file_path.read_text(encoding="utf-8")
    except Exception as e:
        try:
            text = file_path.read_text(encoding="utf-8-sig")
        except Exception:
            raise RuntimeError(f"无法读取文件 {file_path}: {e}") from e
    return [line.strip() for line in text.splitlines() if line.strip()]


def write_result_excel(rows: list[ResultRow], output_path: Path) -> None:
    """将结果写入 Excel，未匹配行标红。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = "匹配结果"
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)
    red_font = Font(color="FF0000")
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if len(row_data) > 4 and row_data[4] not in MATCH_SUCCESS_METHODS:
                cell.font = red_font
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
