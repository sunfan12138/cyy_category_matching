"""将匹配结果写入 Excel，未匹配行标红。"""

from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

ResultRow = tuple[str, str, str, str, str]

HEADERS = ("输入品类", "一级原子品类", "品类编码", "原子品类", "匹配方式")


def write_result_excel(rows: list[ResultRow], output_path: Path) -> None:
    """将结果写入 Excel，未匹配行标红。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = "匹配结果"
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)
    red_font = Font(color="FF0000")
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_data[4] == "未匹配":
                cell.font = red_font
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
