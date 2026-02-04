"""文件读写：从 Excel 读取线索（编码+名称）、匹配结果写入 Excel。"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from core.utils.excel_io import cell_value, open_excel_read

# 匹配成功：关键词匹配到、相似度匹配到、搜索大模型后匹配到；其余为匹配失败（含程序异常）
MATCH_SUCCESS_METHODS = ("规则", "相似度", "搜索后匹配")
# 单条匹配过程中发生未捕获异常时，匹配方式列显示此项，该行会标红
METHOD_EXCEPTION = "程序异常"

# 输入 Excel 列名
INPUT_LEAD_CODE_COL = "线索编码"
INPUT_LEAD_NAME_COL = "线索名称"

# 11 列：线索编码, 输入品类, 一级原子品类, 品类编码, 原子品类, 匹配方式, 品牌, 编码, 原子品类(相似度), 相似度, 大模型描述
ResultRow = tuple[str, str, str, str, str, str, str, str, str, str, str]

# 表头：第一行第 7 列「相似度匹配结果」合并 7～10 列，第 11 列「大模型描述」；其余列合并前两行
HEADER_ROW1_COLS = (
    "线索编码",
    "输入品类",
    "一级原子品类",
    "品类编码",
    "原子品类",
    "匹配方式",
)
HEADER_SIMILARITY_TITLE = "相似度匹配结果"  # 第一行合并 7～10 列
HEADER_ROW2_SIMILARITY = ("品牌", "编码", "原子品类", "相似度")  # 第二行第 7～10 列
HEADER_LLM_DESC = "大模型描述"  # 第一行第 11 列，合并 11:1-2


def _find_column(header: list[str], names: tuple[str, ...]) -> int | None:
    """表头中查找任一名称的列索引（0-based）。"""
    for i, h in enumerate(header):
        if h and str(h).strip() in names:
            return i
    return None


def read_categories_from_file(file_path: Path) -> list[tuple[str, str]]:
    """
    读取 Excel：两列「线索编码」「线索名称」，按线索名称参与匹配。
    返回 list[(线索编码, 线索名称)]，跳过线索名称为空的行。
    """
    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise RuntimeError(f"仅支持 Excel 输入（.xlsx），当前: {path.suffix or path.name}")
    if not path.exists():
        raise RuntimeError(f"文件不存在: {path}")

    with open_excel_read(path) as (_wb, ws):
        if ws is None:
            return []
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row < 2:
            return []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col, values_only=True))
        header = [cell_value(v) for v in (header_row or [])]
        col_code = _find_column(header, (INPUT_LEAD_CODE_COL,))
        col_name = _find_column(header, (INPUT_LEAD_NAME_COL,))
        if col_name is None:
            raise RuntimeError(f"输入 Excel 缺少列「{INPUT_LEAD_NAME_COL}」")
        if col_code is None:
            raise RuntimeError(f"输入 Excel 缺少列「{INPUT_LEAD_CODE_COL}」")
        result: list[tuple[str, str]] = []
        for row_tuple in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            row = list(row_tuple) if row_tuple else []
            name = cell_value(row[col_name]) if col_name < len(row) else ""
            if not name:
                continue
            code = cell_value(row[col_code]) if col_code < len(row) else ""
            result.append((code, name))
        return result


def write_result_excel(rows: list[ResultRow], output_path: Path) -> None:
    """将结果写入 Excel：两行表头（相似度匹配结果拆为 4 子列，其余列合并前两行），未匹配行标红。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = "匹配结果"
    red_font = Font(color="FF0000")

    # 第一行：列 1～6 单列标题，列 7「相似度匹配结果」（合并 7:10），列 11「大模型描述」
    for col, h in enumerate(HEADER_ROW1_COLS, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=1, column=7, value=HEADER_SIMILARITY_TITLE)
    ws.cell(row=1, column=11, value=HEADER_LLM_DESC)
    # 第二行：仅第 7～10 列子标题
    for col, h in enumerate(HEADER_ROW2_SIMILARITY, start=7):
        ws.cell(row=2, column=col, value=h)

    # 合并：前 6 列与最后 1 列合并前两行；第 7 列合并 7～10 列第一行
    for col in (1, 2, 3, 4, 5, 6, 11):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=10)

    # 数据从第 3 行起，11 列
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if len(row_data) > 5 and row_data[5] not in MATCH_SUCCESS_METHODS:
                cell.font = red_font

    wb.save(output_path)


def start_result_excel(output_path: Path) -> None:
    """
    创建结果 Excel 并只写入两行表头（结构同 write_result_excel），保存后关闭。
    用于分块处理时先建文件，再通过 append_result_rows 追加数据行。每次写入后均关闭文件，
    便于程序运行过程中用 Excel 打开查看（避免 Windows 下 EBUSY 占用）。
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")
    ws.title = "匹配结果"
    for col, h in enumerate(HEADER_ROW1_COLS, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=1, column=7, value=HEADER_SIMILARITY_TITLE)
    ws.cell(row=1, column=11, value=HEADER_LLM_DESC)
    for col, h in enumerate(HEADER_ROW2_SIMILARITY, start=7):
        ws.cell(row=2, column=col, value=h)
    for col in (1, 2, 3, 4, 5, 6, 11):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=10)
    wb.save(output_path)
    wb.close()


def append_result_rows(output_path: Path, rows: list[ResultRow]) -> None:
    """
    向已有结果文件追加一批数据行（未匹配/程序异常行标红），保存后关闭工作簿。
    每次只在该函数内打开文件，写完后立即关闭，避免运行中被占用（EBUSY）。
    """
    if not rows:
        return
    output_path = Path(output_path)
    red_font = Font(color="FF0000")
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    if ws is None:
        wb.close()
        raise RuntimeError("无法打开工作表")
    start_row = ws.max_row + 1
    for i, row_data in enumerate(rows):
        row_idx = start_row + i
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if len(row_data) > 5 and row_data[5] not in MATCH_SUCCESS_METHODS:
                cell.font = red_font
    wb.save(output_path)
    wb.close()
